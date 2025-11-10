"""
Módulo de manejo de archivos Excel.

Este módulo contiene funciones para leer y procesar datos desde archivos Excel,
específicamente diseñado para manejar información de contactos y datos de pagos.
"""

import logging
from typing import Dict, List, Optional, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .env_loader import get_excel_path
from .formateo import formato, mayuscula

logger = logging.getLogger(__name__)


def _get_data_name(hoja: Worksheet, row: int) -> Optional[str]:
    """
    Lee el nombre de una fila y devuelve el nombre formateado.
    
    Args:
        hoja: Hoja de Excel activa
        row: Número de fila (1-indexed)
        
    Returns:
        Nombre formateado con formato título, o None si no hay nombre válido
    """
    nombre = hoja.cell(row=row, column=2).value
    if nombre is None or str(nombre).strip() == "":
        return None
    return mayuscula(str(nombre))


def _get_data_telefono(hoja: Worksheet, row: int) -> Optional[str]:
    """
    Lee el teléfono de una fila y devuelve el teléfono formateado.
    
    Args:
        hoja: Hoja de Excel activa
        row: Número de fila (1-indexed)
        
    Returns:
        Teléfono formateado para WhatsApp, o None si no hay teléfono válido
    """
    telefono = hoja.cell(row=row, column=3).value
    if telefono is None or str(telefono).strip() == "":
        return None
    return formato(telefono)


def _is_debe_pagar(hoja: Worksheet, row: int) -> bool:
    """
    Determina si un contacto debe pagar basado en la columna de estado.
    
    Args:
        hoja: Hoja de Excel activa
        row: Número de fila (1-indexed)
        
    Returns:
        True si debe pagar, False si está marcado como "inactiva"
    """
    debe_pagar = hoja.cell(row=row, column=4).value
    if isinstance(debe_pagar, str) and debe_pagar.strip().lower() == "inactiva":
        return False
    return True


def _is_pago_realizado(hoja: Worksheet, row: int, column: int) -> bool:
    """
    Verifica si un pago específico ha sido realizado.
    
    Args:
        hoja: Hoja de Excel activa
        row: Número de fila (1-indexed)
        column: Número de columna (1-indexed)
        
    Returns:
        True si el pago fue realizado, False en caso contrario
    """
    pago_realizado = hoja.cell(row=row, column=column).value
    
    # Verificar diferentes representaciones de "verdadero"
    valores_verdaderos = {"True", "true", "Verdadero", "VERDADERO", True, "SI", "si", "Sí"}
    
    return pago_realizado in valores_verdaderos


def _get_data_length_pago(hoja: Worksheet, row: int = 2) -> int:
    """
    Calcula la cantidad total de columnas de pago disponibles.
    
    Args:
        hoja: Hoja de Excel activa
        row: Fila de referencia para buscar el marcador "Contador" (default: 2)
        
    Returns:
        Número de columnas de pago antes del marcador "Contador"
    """
    i = 5  # Empezar desde la columna E (5)
    length_pago = 0
    
    # Buscar hasta encontrar "Contador" o hasta un límite razonable
    max_columns = 100  # Límite de seguridad
    
    while i <= max_columns:
        valor = hoja.cell(row=row, column=i).value
        if valor == "Contador":
            break
        i += 1
        length_pago += 1
    
    return length_pago


def _get_data_dia_ultimo_pago(hoja: Worksheet, row: int, destino: int) -> Optional[str]:
    """
    Obtiene el día del próximo pago a realizar.
    
    Args:
        hoja: Hoja de Excel activa
        row: Fila de referencia (normalmente 2 para headers)
        destino: Posición del próximo pago
        
    Returns:
        Día del próximo pago o None si no se encuentra
    """
    column_pos = destino + 5  # Ajustar por offset de columnas iniciales
    dia = hoja.cell(row=row, column=column_pos).value
    return str(dia) if dia is not None else None


def _get_data_mes_ultimo_pago(hoja: Worksheet, row: int, dias_pagados: int) -> Optional[str]:
    """
    Obtiene el mes del próximo pago a realizar.
    
    Args:
        hoja: Hoja de Excel activa
        row: Fila de referencia (normalmente 1 para headers de mes)
        dias_pagados: Cantidad de pagos ya realizados
        
    Returns:
        Mes del próximo pago o None si no se encuentra
    """
    i = 5
    mes = hoja.cell(row=row, column=i).value
    
    # Buscar el mes correspondiente al próximo pago
    while i <= (dias_pagados + 5):
        mes_aux = hoja.cell(row=row, column=i).value
        if mes_aux is not None:
            mes = mes_aux
        i += 1
    
    return str(mes) if mes is not None else None


def _get_data_fechas_pago(hoja: Worksheet, row: int) -> Dict[str, Any]:
    """
    Calcula la información completa de pagos para un contacto.
    
    Args:
        hoja: Hoja de Excel activa
        row: Número de fila del contacto (1-indexed)
        
    Returns:
        Diccionario con información de pagos:
        - cantidadPagado: Número de pagos realizados
        - faltantes: Número de pagos pendientes
        - diaAPagar: Día del próximo pago
        - mesAPagar: Mes del próximo pago
    """
    length_pago = _get_data_length_pago(hoja, 2)
    
    i = 5
    cantidad_pagado = 0
    
    # Contar pagos realizados consecutivos
    while i < (5 + length_pago) and _is_pago_realizado(hoja, row, i):
        cantidad_pagado += 1
        i += 1
    
    faltantes = length_pago - cantidad_pagado
    dia_a_pagar = _get_data_dia_ultimo_pago(hoja, 2, cantidad_pagado)
    mes_a_pagar = _get_data_mes_ultimo_pago(hoja, 1, cantidad_pagado)
    
    return {
        "cantidadPagado": cantidad_pagado,
        "faltantes": faltantes,
        "diaAPagar": dia_a_pagar,
        "mesAPagar": mes_a_pagar
    }


def _get_data_row(hoja: Worksheet, row: int) -> Optional[Dict[str, Any]]:
    """
    Procesa una fila completa del Excel y extrae todos los datos del contacto.
    
    Args:
        hoja: Hoja de Excel activa
        row: Número de fila (1-indexed)
        
    Returns:
        Diccionario con datos del contacto o None si debe omitirse
    """
    # Verificar si el contacto debe pagar
    if not _is_debe_pagar(hoja, row):
        logger.debug("Contacto en fila %d marcado como inactivo, omitiendo", row)
        return None
    
    # Obtener datos básicos
    nombre = _get_data_name(hoja, row)
    telefono = _get_data_telefono(hoja, row)
    
    # Validar datos obligatorios
    if nombre is None or telefono is None:
        logger.debug("Contacto en fila %d sin nombre o teléfono válido, omitiendo", row)
        return None
    
    # Obtener datos de pagos
    try:
        data_pagos = _get_data_fechas_pago(hoja, row)
    except Exception as e:
        logger.warning("Error procesando pagos para fila %d: %s", row, e)
        # Datos por defecto si hay error en pagos
        data_pagos = {
            "cantidadPagado": 0,
            "faltantes": 0,
            "diaAPagar": "N/A",
            "mesAPagar": "N/A"
        }
    
    return {
        "nombre": nombre,
        "telefono": telefono,
        "dataPagos": data_pagos
    }


def getData() -> List[Dict[str, Any]]:
    """
    Lee y procesa todos los datos del archivo Excel configurado.
    
    Returns:
        Lista de diccionarios con información de contactos y sus datos de pago.
        Cada diccionario contiene:
        - nombre: Nombre formateado del contacto
        - telefono: Teléfono formateado para WhatsApp
        - dataPagos: Información detallada de pagos
        
    Raises:
        Registra errores en el logger pero no lanza excepciones,
        devuelve lista vacía en caso de errores.
    """
    ruta = get_excel_path()
    logger.info("Usando archivo de Excel: %s", ruta)
    
    # Abrir workbook
    try:
        excel = openpyxl.load_workbook(ruta)
        hoja = excel.active
    except FileNotFoundError:
        logger.error("Archivo no encontrado: %s", ruta)
        return []
    except Exception as e:
        logger.error("Error al abrir el archivo Excel '%s': %s", ruta, e)
        return []
    
    # Verificar que haya suficientes filas
    if hoja.max_row < 3:
        logger.warning("El archivo Excel no tiene suficientes filas de datos")
        return []
    
    lista: List[Dict[str, Any]] = []
    errores_procesamiento = 0
    
    # Procesar cada fila de datos (empezando desde la fila 3)
    for i in range(3, hoja.max_row + 1):
        try:
            fila = _get_data_row(hoja, i)
            if fila is not None and fila.get('dataPagos', {}).get('faltantes', 0) > 0:
                lista.append(fila)
        except Exception as e:
            logger.warning("Error procesando fila %d: %s", i, e)
            errores_procesamiento += 1
    
    logger.info("Procesamiento completado: %d contactos válidos, %d errores",
                len(lista), errores_procesamiento)
    
    return lista
